from openpyxl import Workbook
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Alignment, PatternFill, Font

def createStaffWorksheet(timeRange, staffColumns):
    def contrastTextColor(hexColor):
        r = int(hexColor[0:2],16)
        g = int(hexColor[2:4],16)
        b = int(hexColor[4:6],16)
        if (r * 0.299 + g * 0.587 + b * 0.114) > 186:
            return '000000'
        else:
            return 'FFFFFF'
    def writeTimes(ws, timeRange):
        for row in ws.iter_rows(min_row=2, max_col=1, max_row=len(timeRange.times) + 1):
            for cell in row:
                cell.value = timeRange.times[cell.row - 2]
                cell.number_format = 'HH:MM AM/PM'
    def writeStaffColumns(ws, timeRange, staffColumns):
        for row in ws.iter_cols(min_row=1, min_col=2, max_col=len(staffColumns) + 1, max_row=len(timeRange.times) + 1):
            for cell in row:
                colIndex = column_index_from_string(cell.column)
                staffColIndex = colIndex - 2
                staffCol = staffColumns[staffColIndex]
                currentTime = timeRange.times[cell.row - 2]
                if(cell.row == 1):
                    cell.value = staffCol.name
                    cell.font = Font(bold=True)
                else:
                    for timeSlot in [timeSlot for timeSlot in staffCol.timeSlots if timeSlot.start == currentTime]:
                        cell.value = timeSlot.patient + " - " + timeSlot.info
                        cell.alignment = Alignment(vertical="top",wrap_text=True)
                        cell.fill = PatternFill(start_color=timeSlot.color,
                                end_color=timeSlot.color,
                                fill_type='solid')
                        cell.font = Font(color=contrastTextColor(timeSlot.color))
                        timeRows = int(timeSlot.duration / timeRange.interval)
                        if (timeRows > 1):
                            endRow = cell.row + timeRows - 1
                            ws.merge_cells(
                            start_row = cell.row, end_row = endRow,
                            start_column = colIndex, end_column = colIndex)
    def writeTotalAppts(ws, staffColumns):
        colNum = len(staffColumns) + 2
        labelCell = ws.cell(row=2, column=colNum)
        labelCell.value = "Total Appts:"
        valueCell = ws.cell(row=2, column=colNum + 1)
        valueCell.value = len(staffColumns[-1].timeSlots)
    def columnWidthFix(worksheet):
        for col in worksheet.columns:
         column = col[0].column
         worksheet.column_dimensions[column].width = 14

    wb = Workbook()
    ws = wb.active

    ws.title = "Consult Staff"
    writeTimes(ws, timeRange)
    writeStaffColumns(ws, timeRange, staffColumns)
    writeTotalAppts(ws, staffColumns)
    columnWidthFix(ws)


    wb.save("Staff.xlsx")
